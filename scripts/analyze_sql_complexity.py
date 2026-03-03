#!/usr/bin/env python3
"""
Analyze SQL complexity and generate Excel report.
This script combines two steps:
1. Call sp_analyze_sql to get column/table counts for all queries
2. Generate Excel report with filters and highlights

Usage: python scripts/analyze_sql_complexity.py --min-cols 10 --min-tables 3
"""

import sys
import os
from pathlib import Path
from typing import List, Dict, Any
from collections import defaultdict
import argparse
import json

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.complexity.oracle_complexity_fetcher import OracleComplexityFetcher
from src.utils.logger import setup_logger, create_log_file_path
from src.utils.config_utils import load_env_vars, load_config, get_project_root


def filter_results(
    results: List[Dict[str, Any]],
    min_cols: int = 0,
    min_tables: int = 0
) -> List[Dict[str, Any]]:
    """
    Filter results based on minimum column count or table count.

    Args:
        results: List of complexity results
        min_cols: Minimum column count (0 = no filter)
        min_tables: Minimum table count (0 = no filter)

    Returns:
        Filtered list of results
    """
    if min_cols == 0 and min_tables == 0:
        return results

    filtered = []
    for r in results:
        # Skip errors
        if r.get('error') or r.get('columnCount', -1) < 0 or r.get('tableCount', -1) < 0:
            continue

        # Check if meets criteria
        if r.get('columnCount', 0) > min_cols or r.get('tableCount', 0) > min_tables:
            filtered.append(r)

    return filtered


def create_summary_sheet(ws, results: List[Dict[str, Any]], logger):
    """
    Create summary sheet with complex query count by agent and transaction.

    Sheet structure: Agent | Transaction | Complex Query Count
    """
    logger.info("Creating summary sheet...")

    # Group by agent and transaction
    stats = defaultdict(lambda: {'count': 0})

    for r in results:
        if r.get('error') or r.get('columnCount', -1) < 0 or r.get('tableCount', -1) < 0:
            continue

        key = (r['agent'], r['transaction'])
        stats[key]['count'] += 1

    # Headers
    headers = ['Agent', 'Transaction', 'Complex Query Count']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sort by agent, then transaction
    sorted_keys = sorted(stats.keys(), key=lambda x: (x[0], x[1]))

    # Write data
    row = 2
    for agent, transaction in sorted_keys:
        count = stats[(agent, transaction)]['count']
        ws.cell(row=row, column=1, value=agent)
        ws.cell(row=row, column=2, value=transaction)
        ws.cell(row=row, column=3, value=count)
        row += 1

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    logger.info(f"Summary sheet created with {row - 2} rows")


def create_agent_sheet(wb, agent_name: str, results: List[Dict[str, Any]], min_cols: int, min_tables: int, logger):
    """
    Create a sheet for a specific agent with all queries.

    Columns: Transaction, Query#, Operation, Columns, Tables, SQL Query

    Highlight red: columns > min_cols OR tables > min_tables
    """
    logger.info(f"Creating sheet for agent: {agent_name}")

    # Create sheet with valid name (max 31 chars for Excel)
    sheet_name = agent_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    headers = ['Transaction', 'Query #', 'Operation', 'Columns', 'Tables', 'SQL Query']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sort by transaction, then query number
    sorted_results = sorted(results, key=lambda x: (x['transaction'], x['queryNumber']))

    # Write data
    row = 2
    for r in sorted_results:
        ws.cell(row=row, column=1, value=r['transaction'])
        ws.cell(row=row, column=2, value=r['queryNumber'])
        ws.cell(row=row, column=3, value=r.get('operation', 'UNKNOWN'))

        col_count = r.get('columnCount', -1)
        table_count = r.get('tableCount', -1)

        # Highlight red if exceeds threshold
        col_cell = ws.cell(row=row, column=4, value=col_count)
        table_cell = ws.cell(row=row, column=5, value=table_count)

        # Red highlight for columns
        if min_cols > 0 and col_count > min_cols:
            col_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Red highlight for tables
        if min_tables > 0 and table_count > min_tables:
            table_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Truncate SQL query if too long
        sql = r.get('sqlQuery', '')
        if len(sql) > 10000:
            sql = sql[:10000] + '... (truncated)'
        ws.cell(row=row, column=6, value=sql)

        row += 1

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 100

    # Freeze header row
    ws.freeze_panes = "A2"

    logger.info(f"Agent sheet created with {row - 2} rows")


def generate_excel_report(
    logger,
    results: List[Dict[str, Any]],
    output_file: Path,
    min_cols: int,
    min_tables: int
):
    """
    Generate Excel report from complexity results.

    Args:
        logger: Logger instance
        results: List of complexity results
        output_file: Path to output Excel file
        min_cols: Filter by minimum column count (also used for red highlight)
        min_tables: Filter by minimum table count (also used for red highlight)
    """
    logger.info(f"=" * 60)
    logger.info(f"Generating Excel report for COMPLEX queries...")
    logger.info(f"  Output: {output_file}")
    logger.info(f"  Filter: columns > {min_cols} OR tables > {min_tables}")
    logger.info(f"  Red highlight: columns > {min_cols} OR tables > {min_tables}")
    logger.info(f"  Total records: {len(results)}")

    # Filter results
    filtered = filter_results(results, min_cols, min_tables)
    logger.info(f"  Complex queries found: {len(filtered)} records")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create summary sheet
    summary_ws = wb.create_sheet(title="Summary", index=0)
    create_summary_sheet(summary_ws, filtered, logger)

    # Group by agent
    agent_groups = defaultdict(list)
    for r in filtered:
        agent_groups[r['agent']].append(r)

    # Create sheet for each agent
    for agent_name in sorted(agent_groups.keys()):
        create_agent_sheet(wb, agent_name, agent_groups[agent_name], min_cols, min_tables, logger)

    # Save workbook
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    logger.info(f"=" * 60)
    logger.info(f"Excel report generated successfully!")
    logger.info(f"  File: {output_file}")
    logger.info(f"  Sheets: {len(wb.sheetnames)}")
    logger.info(f"    - Summary")
    for agent in sorted(agent_groups.keys()):
        logger.info(f"    - {agent[:31]}")
    logger.info(f"=" * 60)


def export_jira_csv(
    logger,
    results: List[Dict[str, Any]],
    output_file: Path,
    google_docs_url: str = "https://docs.google.com/spreadsheets/d/PLACEHOLDER_DOC_ID/edit"
):
    """
    Export Jira subtask CSV file.

    CSV format: parent,labels,tên task,mô tả,assignee,reporter,duedate YYYY-MM-DD,estimate (hour)

    Args:
        logger: Logger instance
        results: List of filtered complexity results
        output_file: Path to output CSV file
        google_docs_url: Google Docs URL placeholder
    """
    import csv

    logger.info(f"=" * 60)
    logger.info(f"Exporting Jira subtask CSV...")
    logger.info(f"  Output: {output_file}")

    # Group by agent and transaction
    task_groups = defaultdict(list)
    for r in results:
        if r.get('error') or r.get('columnCount', -1) < 0 or r.get('tableCount', -1) < 0:
            continue
        key = (r['agent'], r['transaction'])
        task_groups[key].append(r)

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)

        # Write header (optional - comment out if header not needed)
        writer.writerow([
            'parent',
            'labels',
            'tên task',
            'mô tả',
            'assignee',
            'reporter',
            'duedate YYYY-MM-DD',
            'estimate (hour)'
        ])

        # Write rows
        for (agent, transaction), queries in sorted(task_groups.items()):
            # Generate label: remove cto- prefix
            agent_label = agent.replace('cto-', '').replace('_', '-')
            labels = f"TOI_UU_SQL,{agent_label}"

            # Task name
            summary = f"Tối ưu API: {transaction}"

            # Count queries
            query_count = len(queries)

            # Description
            description = f'''Tối ưu câu lệnh SQL cho API {transaction}

* Số câu cần tối ưu: {query_count}
* Check các SQL của API {transaction}
* Chi tiết ở sheet: {agent}. Link: {google_docs_url}
* Kiểm tra chỉ truy vấn select những bảng và cột thật sự cần thiết để tối ưu hiệu năng. Nếu nghiệp vụ yêu cầu bắt buộc phải select đủ thì cancel task và comment ghi rõ lý do.

'''

            # Write row
            writer.writerow([
                '',  # parent - placeholder
                labels,  # labels
                summary,  # summary
                description,  # description
                'khoinda',  # assignee
                'nganntk.tgg',  # reporter
                'YYYY-MM-DD',  # duedate - placeholder
                query_count  # estimate (hours)
            ])

    logger.info(f"Jira CSV exported successfully!")
    logger.info(f"  Total tasks: {len(task_groups)}")
    logger.info(f"=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description='Analyze SQL complexity and generate Excel report. '
                    'Queries are filtered and highlighted based on --min-cols and --min-tables.',
        epilog='Example: python analyze_sql_complexity.py --min-cols 10 --min-tables 3'
    )
    parser.add_argument(
        '--min-cols',
        type=int,
        required=True,
        help='Filter queries with column count greater than this value (also used for red highlight)'
    )
    parser.add_argument(
        '--min-tables',
        type=int,
        required=True,
        help='Filter queries with table count greater than this value (also used for red highlight)'
    )
    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output Excel file (default: outputs/complexity/sql_complexity_report.xlsx)'
    )
    parser.add_argument(
        '--csv-output',
        type=str,
        default=None,
        help='Output Jira CSV file (default: outputs/complexity/jira_subtasks.csv)'
    )

    args = parser.parse_args()

    # Load environment variables
    load_env_vars()

    # Load config
    config = load_config()

    # Setup logger
    logs_dir_name = config.get('output', {}).get('logs_dir', 'logs')
    log_file = create_log_file_path(project_root, "complexity_analysis", logs_dir_name)
    logger = setup_logger(
        "SQLComplexityAnalysis",
        log_file=log_file,
        level=config.get('logging', {}).get('level', 'INFO'),
        console_output=config.get('logging', {}).get('console_output', True),
        file_output=True
    )

    logger.info(f"Log file: {log_file}")

    # Get paths from config
    output_config = config.get('output', {})
    base_dir = project_root / output_config.get('base_dir', 'outputs')
    sql_dir = base_dir / output_config.get('sql_dir', 'fetchers/sql')
    sql_info_dir = base_dir / output_config.get('sql_info_dir', 'fetchers/sql_info')
    complexity_output_dir = base_dir / "complexity"

    if args.output:
        excel_output_file = Path(args.output)
    else:
        excel_output_file = complexity_output_dir / "sql_complexity_report.xlsx"

    json_output_file = complexity_output_dir / "sql_complexity_results.json"

    logger.info(f"SQL input directory: {sql_dir}")
    logger.info(f"Excel output file: {excel_output_file}")
    logger.info(f"Filter: columns > {args.min_cols} OR tables > {args.min_tables}")

    # Check if SQL directory exists
    if not sql_dir.exists():
        logger.error(f"SQL directory not found: {sql_dir}")
        logger.info("Run fetch_sql.py first to generate SQL files")
        return 1

    # Initialize fetcher
    fetcher = OracleComplexityFetcher(config)

    try:
        # Step 1: Analyze SQL complexity
        logger.info(f"=" * 60)
        logger.info(f"STEP 1: Analyzing SQL complexity using sp_analyze_sql")
        logger.info(f"=" * 60)

        fetcher.create_pool()
        results = fetcher.analyze_sql_files(
            sql_dir=sql_dir,
            sql_info_dir=sql_info_dir,
            output_file=json_output_file,
            pattern="*.sql"
        )

        logger.info(f"\nStep 1 completed: {len(results)} queries analyzed")

        # Step 2: Generate Excel report
        logger.info(f"\n" + "=" * 60)
        logger.info(f"STEP 2: Generating Excel report")
        logger.info(f"=" * 60)

        generate_excel_report(
            logger=logger,
            results=results,
            output_file=excel_output_file,
            min_cols=args.min_cols,
            min_tables=args.min_tables
        )

        # Step 3: Export Jira CSV
        if args.csv_output:
            csv_output_file = Path(args.csv_output)
        else:
            csv_output_file = complexity_output_dir / "jira_subtasks.csv"

        logger.info(f"\n" + "=" * 60)
        logger.info(f"STEP 3: Exporting Jira subtask CSV")
        logger.info(f"=" * 60)

        # Filter results for CSV export
        filtered_results = filter_results(results, args.min_cols, args.min_tables)

        export_jira_csv(
            logger=logger,
            results=filtered_results,
            output_file=csv_output_file
        )

        logger.info(f"\n" + "=" * 60)
        logger.info(f"✓ SQL complexity analysis completed successfully!")
        logger.info(f"✓ Excel report: {excel_output_file}")
        logger.info(f"✓ JSON results: {json_output_file}")
        logger.info(f"✓ Jira CSV: {csv_output_file}")
        logger.info(f"=" * 60)

        return 0

    except Exception as e:
        logger.error(f"Failed to analyze SQL complexity: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return 1

    finally:
        # Close connection pool
        fetcher.close_pool()


if __name__ == "__main__":
    sys.exit(main())
